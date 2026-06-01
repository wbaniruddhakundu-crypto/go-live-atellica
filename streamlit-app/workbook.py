"""
Atellica Installation Progress workbook support.

Handles uploading the multi-sheet Excel workbook (CI 1900, SCI, Hema, plus
issue logs), parsing every sheet into a clean unified model, persisting the
raw file in Postgres so the shared/published URL keeps the data across
autoscale cold starts, and rendering the viewer pages.
"""

import io
import os
import re
from datetime import datetime, date

import pandas as pd
import plotly.express as px
import streamlit as st

try:
    import psycopg2
except Exception:  # pragma: no cover
    psycopg2 = None


# ── Persistence (Postgres) ──────────────────────────────────────────────────

def _connect():
    if psycopg2 is None:
        raise RuntimeError("psycopg2 is not installed")
    dsn = os.environ.get("DATABASE_URL")
    if not dsn:
        raise RuntimeError("DATABASE_URL is not set")
    return psycopg2.connect(dsn)


def ensure_schema():
    """Create the workbook storage table if it does not exist."""
    try:
        with _connect() as conn, conn.cursor() as cur:
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS atellica_workbook (
                    id SERIAL PRIMARY KEY,
                    filename TEXT,
                    uploaded_at TIMESTAMPTZ DEFAULT now(),
                    content BYTEA NOT NULL
                )
                """
            )
            conn.commit()
        return True
    except Exception:
        return False


def save_workbook(filename: str, content: bytes):
    """Replace the stored workbook with a freshly uploaded one."""
    with _connect() as conn, conn.cursor() as cur:
        cur.execute("DELETE FROM atellica_workbook")
        cur.execute(
            "INSERT INTO atellica_workbook (filename, content) VALUES (%s, %s) RETURNING id",
            (filename, psycopg2.Binary(content)),
        )
        new_id = cur.fetchone()[0]
        conn.commit()
    return new_id


def delete_workbook():
    """Remove the stored workbook so the app returns to an empty state."""
    with _connect() as conn, conn.cursor() as cur:
        cur.execute("DELETE FROM atellica_workbook")
        conn.commit()


def pg_status():
    """Return (ok, error_message) for the Postgres connection."""
    try:
        with _connect() as conn, conn.cursor() as cur:
            cur.execute("SELECT 1")
        return True, ""
    except Exception as e:
        return False, str(e)


def get_workbook_meta():
    """Return (id, filename, uploaded_at) of the latest workbook, or None."""
    try:
        with _connect() as conn, conn.cursor() as cur:
            cur.execute(
                "SELECT id, filename, uploaded_at FROM atellica_workbook ORDER BY id DESC LIMIT 1"
            )
            row = cur.fetchone()
            return (row[0], row[1], row[2]) if row else None
    except Exception:
        return None


def _get_workbook_content(wb_id: int) -> bytes:
    with _connect() as conn, conn.cursor() as cur:
        cur.execute("SELECT content FROM atellica_workbook WHERE id = %s", (wb_id,))
        row = cur.fetchone()
        return bytes(row[0]) if row else b""


# ── Parsing ─────────────────────────────────────────────────────────────────

INSTALL_FIELDS = [
    "instrument_type", "module", "sr_no", "customer_name", "city", "redeployment",
    "project_id", "serial_nos", "sw_ver", "contact",
    "instrument_delivery_date", "ups_ready_date", "water_plant_ready_date",
    "consumables_delivery_date", "svk_kit_delivery_date",
    "installation_start_date", "installation_end_date", "installation_status",
    "srs", "svk_completion", "training_start", "training_completed",
    "extended_validation", "go_live_date",
    "standard_hours", "hours_taken", "standard_time", "time_taken",
    "days_pending", "delay_reason",
    "pending_tasks", "parameters_validated", "challenges_remarks",
]

INSTALL_LABELS = {
    "instrument_type": "Instrument Type", "module": "Module / Config",
    "sr_no": "Sr. No.",
    "customer_name": "Customer", "city": "City / Location",
    "redeployment": "Re-deployment", "project_id": "Project ID",
    "serial_nos": "Serial No(s).", "sw_ver": "SW Ver.", "contact": "Contact",
    "instrument_delivery_date": "Instrument Delivery", "ups_ready_date": "UPS Ready",
    "water_plant_ready_date": "Water Plant Ready",
    "consumables_delivery_date": "Consumables Delivery",
    "svk_kit_delivery_date": "SVK Kit Delivery",
    "installation_start_date": "Install Start", "installation_end_date": "Install End",
    "installation_status": "Install Status", "srs": "SRS",
    "svk_completion": "SVK Completion", "training_start": "Training Start",
    "training_completed": "Training Completed",
    "extended_validation": "Extended Validation", "go_live_date": "Go-Live Date",
    "standard_hours": "Standard Hours",
    "hours_taken": "Hours Taken — Install→Go-Live",
    "standard_time": "Standard Time (days)",
    "time_taken": "Time Taken — Install→Go-Live (days)",
    "days_pending": "Days Pending", "delay_reason": "Reason for Delay in Installation",
    "pending_tasks": "Pending Tasks", "parameters_validated": "Parameters Validated",
    "challenges_remarks": "Challenges / Remarks",
}


def _clean(v):
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v)
    if s.strip().lower() in ("nan", "nat", "none"):
        return ""
    return " ".join(s.split())


def _norm(h) -> str:
    return " ".join(str(h).lower().split()) if h is not None else ""


def _compact(h) -> str:
    return _norm(h).replace(".", "").replace("-", "").replace(" ", "")


def _norm_module(v) -> str:
    """Normalise messy module/config labels so spacing/case variants merge.

    e.g. "DL + IM" / "DL+ IM" -> "DL+IM"; "Atellica SCI" -> "SCI"; "dl+ch" -> "DL+CH".
    """
    s = " ".join(str(v or "").split()).upper()
    if not s or s in ("NA", "N/A", "-", "--", "NONE", "NAN", "NIL"):
        return ""
    s = s.replace(" + ", "+").replace("+ ", "+").replace(" +", "+")
    if s.startswith("ATELLICA "):
        s = s[len("ATELLICA "):].strip()
    return s


def _map_install_header(h):
    n = _norm(h)
    c = _compact(h)
    if not n:
        return None
    # Order matters: specific phrases must be tested before broader substrings
    # (e.g. "reason for the delay in installation start" must not match
    # "installation start", and "customer training start" must not match
    # "customer").
    if c in ("no", "srno", "sno", "h", "slno"):
        return "sr_no"
    if c in ("module", "modules", "config", "configuration", "moduleconfig"):
        return "module"
    if "reason" in n and "delay" in n:
        return "delay_reason"
    if "pending task" in n:
        return "pending_tasks"
    if "days pending" in n:
        return "days_pending"
    if "training" in n and "start" in n:
        return "training_start"
    if "training" in n and ("completed" in n or "complete" in n):
        return "training_completed"
    if "extended validation" in n:
        return "extended_validation"
    if "parameters validated" in n or "parameter validated" in n:
        return "parameters_validated"
    if "standard" in n and ("hour" in n or "hrs" in n or "hr" in c):
        return "standard_hours"
    if ("hour" in n or "hrs" in n) and ("taken" in n or "consum" in n):
        return "hours_taken"
    if "standard" in n or "benchmark" in n:
        return "standard_time"
    if (("time" in n or "days" in n or "day" in n) and ("taken" in n or "consum" in n)) or "actual days" in n:
        return "time_taken"
    if "customer" in n:
        return "customer_name"
    if "re-deployment" in n or "redeployment" in n or "re deployment" in n:
        return "redeployment"
    if "city" in n or "location" in n:
        return "city"
    if "project id" in n:
        return "project_id"
    if "serial" in n:
        return "serial_nos"
    if "system software" in n or "sw ver" in n or "software version" in n or "sw.ver" in n:
        return "sw_ver"
    if "contact" in n or "e-mail" in n or "email" in n:
        return "contact"
    if "instrument delivery" in n:
        return "instrument_delivery_date"
    if "ups ready" in n:
        return "ups_ready_date"
    if "water plant" in n:
        return "water_plant_ready_date"
    if "consum" in n:
        return "consumables_delivery_date"
    if "svk kit" in n:
        return "svk_kit_delivery_date"
    if "svk completion" in n:
        return "svk_completion"
    if "installation start" in n:
        return "installation_start_date"
    if "installation end" in n:
        return "installation_end_date"
    if "install" in n and "status" in n:
        return "installation_status"
    if "go live" in n or "golive" in c:
        return "go_live_date"
    if "challn" in n or "challenge" in n or "remark" in n:
        return "challenges_remarks"
    if n == "srs" or c == "srs":
        return "srs"
    return None


def _map_aux_header(h):
    n = _norm(h)
    c = _compact(h)
    if c in ("no", "srno", "sno"):
        return "sr_no"
    if "customer" in n:
        return "customer_name"
    if "instrument" in n:
        return "instrument"
    if "module" in n:
        return "module"
    if "corrective" in n or "solution" in n or "action" in n:
        return "resolution"
    if "part replaced" in n or n == "part" or "part" in n:
        return "part_replaced"
    if "problem" in n or "issue" in n or "faced" in n:
        return "problem"
    return None


def _sheet_kind(name: str):
    n = _norm(name)
    if "app" in n and "issue" in n:
        return ("app_issues", None)
    if "hardware" in n:
        return ("hardware_issues", None)
    if "part" in n and "fail" in n:
        return ("part_failures", None)
    if "hema" in n:
        return ("install", "Hema")
    if ("ci 1900" in n or "ci1900" in n) and "neigh" in n:
        return ("install", "CI 1900 (Neighbouring)")
    if "ci 1900" in n or "ci1900" in n:
        return ("install", "CI 1900 (India)")
    if "sci" in n:
        return ("install", "SCI")
    return (None, None)


def _find_header(rows, min_cols):
    for idx, r in enumerate(rows[:8]):
        strs = [c for c in r if isinstance(c, str) and c.strip()]
        if len(strs) >= min_cols:
            return idx
    return None


@st.cache_data(show_spinner=False)
def parse_workbook(wb_id: int):
    """Parse the stored workbook into category DataFrames. Cached per workbook id."""
    import openpyxl

    content = _get_workbook_content(wb_id)
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    install_records = []
    app_issues, hardware_issues, part_failures = [], [], []

    for ws in wb.worksheets:
        kind, inst_type = _sheet_kind(ws.title)
        if kind is None:
            continue
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        if kind == "install":
            hdr = _find_header(rows, 4)
            if hdr is None:
                continue
            colmap = {}
            for ci, cell in enumerate(rows[hdr]):
                key = _map_install_header(cell)
                if key and key not in colmap.values():
                    colmap[ci] = key
            # Some sheets (e.g. SCI) carry the analyzer module/config in an
            # unlabeled column just before "Serial Nos." — capture it.
            if "module" not in colmap.values():
                serial_ci = next(
                    (ci for ci, k in colmap.items() if k == "serial_nos"), None
                )
                if serial_ci is not None and serial_ci - 1 >= 0 and (serial_ci - 1) not in colmap:
                    hcell = rows[hdr][serial_ci - 1] if serial_ci - 1 < len(rows[hdr]) else None
                    if not (isinstance(hcell, str) and hcell.strip()):
                        colmap[serial_ci - 1] = "module"
            # Some sheets (e.g. CI 1900) hold the actual delay-reason / remark
            # text in unlabeled columns AFTER the last headed column. Capture
            # those trailing columns so the notes aren't silently dropped.
            # Scan only up to the last column that actually holds data (the raw
            # row tuple can span Excel's full 16k width), not the nominal width.
            last_mapped = max(colmap.keys()) if colmap else -1
            last_nonempty = last_mapped
            for r in rows[hdr + 1:]:
                for ci in range(len(r) - 1, last_nonempty, -1):
                    if _clean(r[ci]):
                        last_nonempty = ci
                        break
            extra_cols = [ci for ci in range(last_mapped + 1, last_nonempty + 1) if ci not in colmap]
            for r in rows[hdr + 1:]:
                rec = {f: "" for f in INSTALL_FIELDS}
                rec["instrument_type"] = inst_type
                for ci, key in colmap.items():
                    if ci < len(r):
                        rec[key] = _clean(r[ci])
                extras = []
                for ci in extra_cols:
                    if ci < len(r):
                        val = _clean(r[ci])
                        if val:
                            extras.append(val)
                if extras:
                    rec["challenges_remarks"] = " | ".join(
                        x for x in [rec.get("challenges_remarks", "").strip(), *extras] if x
                    )
                if rec["customer_name"]:
                    install_records.append(rec)
        else:
            hdr = _find_header(rows, 3)
            if hdr is None:
                continue
            colmap = {}
            for ci, cell in enumerate(rows[hdr]):
                key = _map_aux_header(cell)
                if key and key not in colmap.values():
                    colmap[ci] = key
            for r in rows[hdr + 1:]:
                rec = {}
                for ci, key in colmap.items():
                    if ci < len(r):
                        rec[key] = _clean(r[ci])
                if not rec.get("customer_name"):
                    continue
                if kind == "app_issues":
                    app_issues.append({
                        "sr_no": rec.get("sr_no", ""),
                        "customer_name": rec.get("customer_name", ""),
                        "problem": rec.get("problem", ""),
                        "solution": rec.get("resolution", ""),
                    })
                elif kind == "hardware_issues":
                    hardware_issues.append({
                        "sr_no": rec.get("sr_no", ""),
                        "customer_name": rec.get("customer_name", ""),
                        "instrument": rec.get("instrument", ""),
                        "module": rec.get("module", ""),
                        "problem": rec.get("problem", ""),
                        "corrective_action": rec.get("resolution", ""),
                    })
                elif kind == "part_failures":
                    part_failures.append({
                        "sr_no": rec.get("sr_no", ""),
                        "customer_name": rec.get("customer_name", ""),
                        "problem": rec.get("problem", ""),
                        "part_replaced": rec.get("part_replaced", ""),
                    })

    install_df = pd.DataFrame(install_records, columns=INSTALL_FIELDS) if install_records else pd.DataFrame(columns=INSTALL_FIELDS)
    return {
        "installations": install_df,
        "app_issues": pd.DataFrame(app_issues) if app_issues else pd.DataFrame(columns=["sr_no", "customer_name", "problem", "solution"]),
        "hardware_issues": pd.DataFrame(hardware_issues) if hardware_issues else pd.DataFrame(columns=["sr_no", "customer_name", "instrument", "module", "problem", "corrective_action"]),
        "part_failures": pd.DataFrame(part_failures) if part_failures else pd.DataFrame(columns=["sr_no", "customer_name", "problem", "part_replaced"]),
    }


# ── Derived helpers ─────────────────────────────────────────────────────────

def _is_completed(status: str) -> bool:
    return "complete" in str(status).lower()


def _has_delay(row) -> bool:
    if str(row.get("delay_reason", "")).strip():
        return True
    dp = str(row.get("days_pending", "")).strip().lower()
    if dp and dp not in ("0", "nil", "na", "no", "-"):
        return True
    return False


# ── Timing: install → go-live vs standard ───────────────────────────────────

def _to_num(value):
    """Pull the first number out of a cell (handles '15 days', '20', etc.)."""
    if value is None:
        return None
    m = re.search(r"-?\d+(?:\.\d+)?", str(value).replace(",", ""))
    return float(m.group()) if m else None


def _to_date(value):
    s = str(value).strip()
    if not s or s.lower() in ("nan", "nat", "none", "-"):
        return None
    d = pd.to_datetime(s, dayfirst=True, errors="coerce")
    return d if pd.notna(d) else None


def _actual_days(row):
    """Actual install→go-live days: explicit column if given, else from dates.

    Guards against bad date cells (e.g. a typo'd year) by rejecting negative or
    absurdly large durations so they don't pollute the comparison chart.
    """
    tt = _to_num(row.get("time_taken", ""))
    if tt is not None:
        return tt if 0 <= tt <= 3650 else None
    gl = _to_date(row.get("go_live_date", ""))
    start = _to_date(row.get("installation_start_date", "")) or _to_date(
        row.get("installation_end_date", "")
    )
    if gl is not None and start is not None:
        delta = (gl - start).days
        return delta if 0 <= delta <= 3650 else None
    return None


def _row_actual(row, mode: str):
    if mode == "hours":
        return _to_num(row.get("hours_taken", ""))
    return _actual_days(row)


def _timing_status(row, mode: str = "hours", std_override=None):
    if std_override is not None:
        std = std_override
    elif mode == "hours":
        std = _to_num(row.get("standard_hours", ""))
    else:
        std = _to_num(row.get("standard_time", ""))
    actual = _row_actual(row, mode)
    if std is None or actual is None:
        return "Not enough data"
    return "Exceeded standard" if actual > std else "Within standard"


def _exceeded_amount(row, mode: str = "hours", std_override=None):
    """How much an instrument went over its standard (>=0), or None if unknown."""
    if std_override is not None:
        std = std_override
    elif mode == "hours":
        std = _to_num(row.get("standard_hours", ""))
    else:
        std = _to_num(row.get("standard_time", ""))
    actual = _row_actual(row, mode)
    if std is None or actual is None:
        return None
    return max(0.0, actual - std)


def render_timing_pie(df, key_prefix: str):
    """Pie of instruments within (green) vs over (red) the standard time.

    Two options: compare by Hours (Standard Hours vs Hours Taken) or by Days
    (Standard Time vs actual days from install→go-live). When the Excel has no
    per-instrument standard column, the user can type a single standard value
    that applies to every instrument so the comparison still works.
    """
    if df is None or df.empty:
        return
    unit = st.radio(
        "Compare by", ["Hours", "Days"], horizontal=True,
        key=f"{key_prefix}_timing_unit",
    )
    mode = unit.lower()
    word = "hours" if mode == "hours" else "days"

    has_actual = df.apply(lambda r: _row_actual(r, mode) is not None, axis=1).any()
    if not has_actual:
        if mode == "hours":
            st.info(
                "💡 To compare by **hours**, add an **Hours Taken** column to your Excel "
                "(the actual hours each instrument took). Optionally add a **Standard "
                "Hours** column too, or just type a standard value here once the data is in."
            )
        else:
            st.info(
                "💡 To compare by **days**, the app needs each instrument's **Installation "
                "Start Date** and **Go-Live Date** (or a **Time Taken** column). None were "
                "found in the current workbook."
            )
        return

    std_field = "standard_hours" if mode == "hours" else "standard_time"
    has_row_std = std_field in df.columns and df[std_field].map(
        lambda v: _to_num(v) is not None
    ).any()

    std_override = None
    if not has_row_std:
        st.caption(
            f"No per-instrument standard {word} found in the Excel — set one below "
            f"to compare against (or add a **Standard {word.title()}** column)."
        )
        val = st.number_input(
            f"Standard {word} (applies to all instruments)",
            min_value=0.0, value=0.0, step=1.0,
            key=f"{key_prefix}_global_std_{mode}",
        )
        std_override = val if val and val > 0 else None
        if std_override is None:
            return

    status = df.apply(lambda r: _timing_status(r, mode, std_override), axis=1)
    if not (status != "Not enough data").any():
        st.info("Not enough data to build the comparison chart.")
        return

    color_map = {
        "Within standard": "#22c55e",
        "Exceeded standard": "#ef4444",
        "Not enough data": "#94a3b8",
    }
    counts = status.value_counts().reset_index()
    counts.columns = ["Timing", "Count"]
    fig = px.pie(
        counts, names="Timing", values="Count", hole=0.45,
        color="Timing", color_discrete_map=color_map,
    )
    fig.update_traces(textinfo="value+percent")
    fig.update_layout(height=340, margin=dict(t=10, b=10, l=10, r=10))
    st.plotly_chart(fig, width="stretch", key=f"{key_prefix}_timing_pie")

    exceeded = int((status == "Exceeded standard").sum())
    within = int((status == "Within standard").sum())
    unknown = int((status == "Not enough data").sum())
    msg = (
        f"🔴 {exceeded} took longer than standard · 🟢 {within} within standard"
    )
    if unknown:
        msg += f" · ⚪ {unknown} missing date/standard data"
    st.caption(msg)

    # ── Extra breakdowns: total time exceeded + money gap ───────────────────
    over = df.copy()
    over["_exceeded"] = df.apply(
        lambda r: _exceeded_amount(r, mode, std_override), axis=1
    )
    over = over[(over["_exceeded"].notna()) & (over["_exceeded"] > 0)]
    if over.empty:
        return

    # Group exceeded amounts by module/configuration so each module with data
    # (DL+IM, SH+IM, SCI, …) shows separately instead of collapsing into one
    # instrument-type slice. Rows without a module fall back to their type.
    def _group_label(r):
        mod = _norm_module(r.get("module", "")) if "module" in over.columns else ""
        if mod:
            return mod
        t = str(r.get("instrument_type", "") or "").strip()
        return t if t and t.lower() != "nan" else "Unknown"

    over["_group"] = over.apply(_group_label, axis=1)
    by_type = (
        over.groupby("_group")["_exceeded"].sum().reset_index()
        .rename(columns={"_exceeded": "Exceeded", "_group": "Module"})
        .sort_values("Exceeded", ascending=False)
    )

    st.markdown(f"**Total {word} exceeded — by module / configuration**")
    fig_t = px.pie(by_type, names="Module", values="Exceeded", hole=0.4)
    fig_t.update_traces(textinfo="value+percent")
    fig_t.update_layout(height=340, margin=dict(t=10, b=10, l=10, r=10))
    st.plotly_chart(fig_t, width="stretch", key=f"{key_prefix}_exceeded_pie")
    st.caption(f"Total {word} over standard across all instruments: {by_type['Exceeded'].sum():,.0f}")

    if mode == "hours":
        rate = st.number_input(
            "Cost per exceeded hour (INR)",
            min_value=0.0, value=3000.0, step=500.0,
            key=f"{key_prefix}_hour_rate",
        )
        by_type["Cost"] = by_type["Exceeded"] * rate
        total_cost = float(by_type["Cost"].sum())
        st.markdown(f"**Money gap from exceeded hours — by module / configuration** (total ₹{total_cost:,.0f})")
        fig_c = px.pie(by_type, names="Module", values="Cost", hole=0.4)
        fig_c.update_traces(
            textinfo="label+percent",
            hovertemplate="%{label}: ₹%{value:,.0f} (%{percent})<extra></extra>",
        )
        fig_c.update_layout(height=340, margin=dict(t=10, b=10, l=10, r=10))
        st.plotly_chart(fig_c, width="stretch", key=f"{key_prefix}_cost_pie")
        st.caption(
            f"At ₹{rate:,.0f}/hour, the exceeded hours add up to ₹{total_cost:,.0f} in extra cost."
        )


# ── Rendering: Upload ───────────────────────────────────────────────────────

def _db_error_banner():
    ok, err = pg_status()
    if not ok:
        st.error(
            "The data store is currently unreachable, so workbooks cannot be loaded or "
            "saved right now. Please try again in a moment.\n\n"
            f"_Technical detail: {err}_"
        )
    return ok


def render_upload():
    st.title("📤 Upload Installation Workbook")
    st.caption(
        "Upload your **Atellica Installation Progress** Excel file. The app reads every "
        "sheet — CI 1900, SCI, Hema and the issue logs — and keeps it on the shared link."
    )

    if not _db_error_banner():
        return

    meta = get_workbook_meta()
    if meta:
        _id, fname, uploaded_at = meta
        when = uploaded_at.strftime("%d %b %Y, %H:%M") if hasattr(uploaded_at, "strftime") else str(uploaded_at)
        st.success(f"Currently loaded: **{fname or 'workbook.xlsx'}**  ·  uploaded {when}")
        try:
            data = parse_workbook(_id)
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Instruments", len(data["installations"]))
            c2.metric("App Issues", len(data["app_issues"]))
            c3.metric("Hardware Issues", len(data["hardware_issues"]))
            c4.metric("Part Failures", len(data["part_failures"]))
            st.divider()
            _download_everything(meta, data)
        except Exception as e:
            st.warning(f"Could not build the summary / backup downloads: {e}")

        with st.expander("🗑️ Delete the current workbook"):
            st.caption(
                "This removes the stored file so the app goes back to empty. "
                "You can upload a new workbook any time afterwards."
            )
            confirm = st.checkbox(
                "Yes, I want to delete the current workbook", key="wb_delete_confirm"
            )
            if st.button("Delete workbook", type="secondary", disabled=not confirm):
                try:
                    delete_workbook()
                    parse_workbook.clear()
                except Exception as e:
                    st.error(f"Could not delete the workbook: {e}")
                else:
                    st.success("Workbook deleted. You can upload a new one below.")
                    st.rerun()

    st.divider()
    uploaded = st.file_uploader(
        "Choose your Excel workbook (.xlsx)",
        type=["xlsx"],
        help="The same file you maintain in Excel — all sheets are imported automatically.",
    )

    if uploaded is not None:
        content = uploaded.getvalue()
        try:
            import openpyxl
            openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        except Exception as e:
            st.error(f"That file could not be read as an Excel workbook: {e}")
            return

        if st.button("💾 Save & import this workbook", type="primary"):
            try:
                new_id = save_workbook(uploaded.name, content)
                parse_workbook.clear()
                data = parse_workbook(new_id)
            except Exception as e:
                st.error(f"Could not save the workbook: {e}")
                return
            st.success("Workbook imported and saved. It is now live on your shared link.")
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Instruments", len(data["installations"]))
            c2.metric("App Issues", len(data["app_issues"]))
            c3.metric("Hardware Issues", len(data["hardware_issues"]))
            c4.metric("Part Failures", len(data["part_failures"]))
            if not data["installations"].empty:
                st.markdown("**Instruments by type**")
                counts = data["installations"]["instrument_type"].value_counts().reset_index()
                counts.columns = ["Instrument Type", "Count"]
                st.dataframe(counts, width="stretch", hide_index=True)
            st.info("Open **🔧 Installation Tracker** in the sidebar to explore the data.")


# ── Rendering: Installation Tracker ─────────────────────────────────────────

def _load_current():
    meta = get_workbook_meta()
    if not meta:
        return None, None
    return meta, parse_workbook(meta[0])


def render_tracker():
    st.title("🔧 Installation Tracker")
    if not _db_error_banner():
        return
    meta, data = _load_current()
    if not data:
        st.info("No workbook loaded yet. Go to **📤 Upload Workbook** to add your Excel file.")
        return

    df = data["installations"].copy()
    if df.empty:
        st.warning("No installation rows were found in the uploaded workbook.")
        return

    st.caption("All instruments across CI 1900, SCI and Hema — with delay reasons and go-live dates.")

    # ── Filters
    f1, f2, f3 = st.columns([1.2, 1.2, 2])
    types = sorted(df["instrument_type"].dropna().unique().tolist())
    sel_types = f1.multiselect("Instrument type", types, default=types)
    only_delayed = f2.checkbox("Only show delayed", value=False)
    query = f3.text_input("Search customer / city / serial / reason", "")

    view = df[df["instrument_type"].isin(sel_types)] if sel_types else df.iloc[0:0]
    if only_delayed:
        view = view[view.apply(_has_delay, axis=1)]
    if query.strip():
        q = query.strip().lower()
        searchcols = ["customer_name", "city", "serial_nos", "delay_reason", "pending_tasks", "challenges_remarks"]
        mask = view[searchcols].apply(lambda r: q in " ".join(r.astype(str)).lower(), axis=1)
        view = view[mask]

    # ── KPIs
    total = len(view)
    completed = int(view["installation_status"].apply(_is_completed).sum())
    delayed = int(view.apply(_has_delay, axis=1).sum())
    in_progress = total - completed
    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Instruments", total)
    k2.metric("Installation Completed", completed)
    k3.metric("In Progress / Pending", in_progress)
    k4.metric("Delayed", delayed, delta=f"-{delayed}" if delayed else None, delta_color="inverse")

    # ── Charts
    import plotly.express as px
    cc1, cc2 = st.columns(2)
    with cc1:
        st.markdown("**By instrument type**")
        tc = view["instrument_type"].value_counts().reset_index()
        tc.columns = ["Instrument Type", "Count"]
        if not tc.empty:
            fig = px.bar(tc, x="Instrument Type", y="Count", color="Instrument Type", text="Count")
            fig.update_layout(showlegend=False, height=320, margin=dict(t=10, b=10))
            st.plotly_chart(fig, width="stretch")
    with cc2:
        st.markdown("**By installation status**")
        sc = view["installation_status"].replace("", "Not set").value_counts().reset_index()
        sc.columns = ["Status", "Count"]
        if not sc.empty:
            fig2 = px.pie(sc, names="Status", values="Count", hole=0.45)
            fig2.update_layout(height=320, margin=dict(t=10, b=10))
            st.plotly_chart(fig2, width="stretch")

    mview = view.copy()
    mview["_module"] = mview["module"].map(_norm_module)
    mview = mview[mview["_module"] != ""]
    if not mview.empty:
        mc = mview["_module"].value_counts().reset_index()
        mc.columns = ["Module", "Count"]
        mod_order = mc["Module"].tolist()  # shared descending-count order

        st.markdown("**By module / configuration**")
        figm = px.bar(
            mc, x="Module", y="Count", color="Module", text="Count",
            category_orders={"Module": mod_order},
        )
        figm.update_layout(showlegend=False, height=360, margin=dict(t=10, b=10))
        st.plotly_chart(figm, width="stretch")

        st.markdown("**Installation status — by module / configuration**")
        mview["_status"] = mview["installation_status"].apply(
            lambda s: "Completed" if _is_completed(s) else "In progress / Pending"
        )
        sm = (
            mview.groupby(["_module", "_status"]).size().reset_index(name="Count")
            .rename(columns={"_module": "Module", "_status": "Status"})
        )
        figsm = px.bar(
            sm, x="Module", y="Count", color="Status", text="Count", barmode="stack",
            category_orders={"Module": mod_order},
            color_discrete_map={"Completed": "#22c55e", "In progress / Pending": "#f59e0b"},
        )
        figsm.update_layout(height=360, margin=dict(t=10, b=10))
        st.plotly_chart(figsm, width="stretch")

    st.markdown("**Installation → Go-Live time vs standard**")
    render_timing_pie(view, "tracker")

    st.divider()

    tab1, tab2 = st.tabs(["📋 All Installations", "⏳ Delays & Go-Live"])

    with tab1:
        show_all = st.checkbox("Show every column from the Excel", value=False)
        if show_all:
            cols = INSTALL_FIELDS
        else:
            cols = [
                "instrument_type", "module", "customer_name", "city", "serial_nos",
                "installation_status", "go_live_date", "days_pending", "delay_reason",
            ]
        out = view[cols].rename(columns=INSTALL_LABELS)
        st.dataframe(out, width="stretch", hide_index=True)
        _download_buttons(view.rename(columns=INSTALL_LABELS), "atellica_installations")

    with tab2:
        delay_view = view[view.apply(
            lambda r: bool(str(r.get("delay_reason", "")).strip()
                           or str(r.get("pending_tasks", "")).strip()
                           or str(r.get("days_pending", "")).strip()), axis=1)]
        st.caption(f"{len(delay_view)} instrument(s) with a pending task, delay reason or days pending.")
        cols = [
            "instrument_type", "customer_name", "city", "go_live_date",
            "days_pending", "delay_reason", "pending_tasks", "challenges_remarks",
        ]
        out = delay_view[cols].rename(columns=INSTALL_LABELS)
        st.dataframe(out, width="stretch", hide_index=True)
        _download_buttons(out, "atellica_delays")


def _download_everything(meta, data):
    """Offline backup: one combined Excel of all parsed data + the original file.

    Lets the user keep a full copy of everything on their own computer, so the
    data survives even if the published link is taken offline.
    """
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    st.markdown("### 💾 Download everything (offline backup)")
    st.caption(
        "Save a full copy to your own computer. This works even if the shared "
        "link is later switched off — your data lives in these files."
    )

    sheets = {
        "Instruments": data["installations"].rename(columns=INSTALL_LABELS),
        "App Issues": data["app_issues"],
        "Hardware Issues": data["hardware_issues"],
        "Part Failures": data["part_failures"],
    }
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for name, frame in sheets.items():
            out = frame if frame is not None and not frame.empty else pd.DataFrame()
            # Excel sheet names are capped at 31 chars and must be non-empty.
            out.to_excel(writer, index=False, sheet_name=name[:31])

    c1, c2 = st.columns(2)
    c1.download_button(
        "⬇️ Download all data (Excel)",
        data=buf.getvalue(),
        file_name=f"atellica_all_data_{ts}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        width="stretch",
        type="primary",
    )

    raw = None
    try:
        raw = _get_workbook_content(meta[0])
    except Exception:
        raw = None
    orig_name = meta[1] or "Atellica_Installation_Progress.xlsx"
    c2.download_button(
        "⬇️ Download original uploaded file",
        data=raw if raw else b"",
        file_name=orig_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        width="stretch",
        disabled=not raw,
    )


def _download_buttons(df: pd.DataFrame, basename: str):
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    c1, c2 = st.columns(2)
    c1.download_button(
        "⬇️ Download CSV", data=df.to_csv(index=False).encode("utf-8"),
        file_name=f"{basename}_{ts}.csv", mime="text/csv", width="stretch",
    )
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Data")
    c2.download_button(
        "⬇️ Download Excel", data=buf.getvalue(),
        file_name=f"{basename}_{ts}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        width="stretch",
    )


# ── Rendering: section embedded in the legacy Project List page ──────────────

def render_project_list_section():
    """Compact installation-workbook table for embedding in the Project List page."""
    if not pg_status()[0]:
        return
    meta = get_workbook_meta()
    if not meta:
        st.subheader("🔧 Installation Workbook Instruments")
        st.info("No installation workbook uploaded yet — see **📤 Upload Workbook**.")
        return
    df = parse_workbook(meta[0])["installations"]
    st.subheader(f"🔧 Installation Workbook Instruments ({len(df)})")
    if df.empty:
        st.caption("No instruments found in the uploaded workbook.")
        return

    f1, f2 = st.columns([1.4, 2])
    types = sorted(df["instrument_type"].dropna().unique().tolist())
    sel = f1.multiselect("Instrument type", types, default=types, key="pl_wb_types")
    q = f2.text_input("Search customer / city / serial / reason", "", key="pl_wb_search")

    view = df[df["instrument_type"].isin(sel)] if sel else df.iloc[0:0]
    if q.strip():
        ql = q.strip().lower()
        scols = ["customer_name", "city", "serial_nos", "delay_reason"]
        view = view[view[scols].apply(lambda r: ql in " ".join(r.astype(str)).lower(), axis=1)]

    st.caption(f"Showing {len(view)} of {len(df)} instruments")

    mods = view["module"].map(_norm_module)
    mods = mods[mods != ""]
    if not mods.empty:
        import plotly.express as px
        st.markdown("**By module / configuration**")
        mc = mods.value_counts().reset_index()
        mc.columns = ["Module", "Count"]
        figm = px.bar(mc, x="Module", y="Count", color="Module", text="Count")
        figm.update_layout(showlegend=False, height=360, margin=dict(t=10, b=10))
        st.plotly_chart(figm, width="stretch")

    st.markdown("**Installation → Go-Live time vs standard**")
    render_timing_pie(view, "pl")

    show_all = st.checkbox("Show every column from the Excel", value=True, key="pl_wb_showall")
    cols = INSTALL_FIELDS if show_all else [
        "instrument_type", "module", "customer_name", "city", "serial_nos",
        "installation_status", "go_live_date", "days_pending", "delay_reason",
    ]
    out = view[cols].rename(columns=INSTALL_LABELS)
    st.dataframe(out, width="stretch", hide_index=True)
    _download_buttons(out, "atellica_project_list")


# ── Rendering: Issues & Failures ────────────────────────────────────────────

def render_issues():
    st.title("⚠️ Issues & Failures Log")
    if not _db_error_banner():
        return
    meta, data = _load_current()
    if not data:
        st.info("No workbook loaded yet. Go to **📤 Upload Workbook** to add your Excel file.")
        return

    app_df = data["app_issues"]
    hw_df = data["hardware_issues"]
    pf_df = data["part_failures"]

    t1, t2, t3 = st.tabs([
        f"🖥️ Application Issues ({len(app_df)})",
        f"🔩 Hardware Issues ({len(hw_df)})",
        f"🧰 Part Failures ({len(pf_df)})",
    ])

    with t1:
        _issue_table(app_df, {
            "sr_no": "No.", "customer_name": "Customer",
            "problem": "Problem Faced", "solution": "Solution",
        }, "Search application issues")
    with t2:
        _issue_table(hw_df, {
            "sr_no": "No.", "customer_name": "Customer", "instrument": "Instrument",
            "module": "Module", "problem": "Problem Description",
            "corrective_action": "Corrective Action",
        }, "Search hardware issues")
    with t3:
        _issue_table(pf_df, {
            "sr_no": "No.", "customer_name": "Customer",
            "problem": "Problem Faced", "part_replaced": "Part Replaced",
        }, "Search part failures")


def _issue_table(df: pd.DataFrame, labels: dict, search_label: str):
    if df.empty:
        st.info("No rows for this category in the uploaded workbook.")
        return
    q = st.text_input(search_label, "", key=search_label)
    view = df
    if q.strip():
        ql = q.strip().lower()
        view = df[df.apply(lambda r: ql in " ".join(r.astype(str)).lower(), axis=1)]
    out = view[[c for c in labels if c in view.columns]].rename(columns=labels)
    st.dataframe(out, width="stretch", hide_index=True)
